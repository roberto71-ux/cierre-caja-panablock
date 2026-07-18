#!/usr/bin/env python3
"""
GRUPO PANABLOCK — Cierre Diario de Caja — Módulo central
Sin dependencias de GUI. Importado por:
  • app.py        (interfaz Streamlit / web)
  • cierre_caja.py (interfaz tkinter / escritorio)
"""

import re
import io
import os
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
try:
    from openpyxl.drawing.image import Image as XLImage
    _HAS_XL_IMAGE = True
except ImportError:
    _HAS_XL_IMAGE = False

# ════════════════════════════════════════════════════════════════════
# ESTILOS EXCEL
# ════════════════════════════════════════════════════════════════════
DARK_BLUE  = '1F4E79'
LIGHT_BLUE = 'BDD7EE'
LGRAY      = 'D9D9D9'
MGRAY      = 'BFBFBF'
RED_FILL   = 'FF0000'
WHITE      = 'FFFFFF'
CURRENCY   = '#,##0.00'

_thin = Side(border_style='thin', color='000000')
BRD   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CA    = Alignment(horizontal='center', vertical='center')
LA    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
LA2   = Alignment(horizontal='left',   vertical='center', wrap_text=True, indent=2)
RA    = Alignment(horizontal='right',  vertical='center')

def mk_fill(hex_c):
    return PatternFill(start_color=hex_c, end_color=hex_c, fill_type='solid')

NO_BORDER = Border()

# ════════════════════════════════════════════════════════════════════
# CLASIFICADOR DE GASTOS
# ════════════════════════════════════════════════════════════════════

# Mapa de proveedores cargado desde proveedores.xlsx.
# Se asigna desde fuera: core._PROV_MAP = {...}
# Clave: keyword en MAYÚSCULAS. Valor: categoría.
_PROV_MAP = {}

# Keywords genéricas como segunda línea de defensa
GASTO_KEYWORDS = [
    ('Materia Prima',           ['cemento', 'arena', 'agregado', 'bloque',
                                  'hormigon', 'hormigón', 'material prima']),
    ('Pigmento y Aditivos',     ['pigmento', 'aditivo', 'colorante',
                                  'oxido', 'óxido']),
    ('Mantenimiento de Planta', ['mantenimiento', 'reparacion', 'reparación',
                                  'repuesto', 'herramienta']),
    ('Servicios Básicos',       ['agua', 'luz', 'electricidad', 'energia',
                                  'energía', 'internet', 'telefono', 'teléfono',
                                  'cable', 'gas', 'assa', 'naturgy',
                                  'claro', 'tigo', 'cwp']),
    ('Tarimas',                 ['tarima', 'pallet']),
    ('Transporte',              ['transporte', 'flete', 'envio', 'envío',
                                  'courier', 'acarreo', 'traslado']),
    ('Acreedores',              ['proveedor', 'acreedor', 'pago a proveedor']),
]

ITEMS_PAGOS = [
    'Acreedores', 'Mantenimiento de Planta', 'Materia Prima',
    'Pigmento y Aditivos', 'Servicios Básicos', 'Tarimas',
    'Transporte', 'Otros',
]

FORMAS_ORDEN = ['Efectivo', 'ACH / Transferencia', 'Yappy', 'Tarjeta', 'Otros']

def _classify_gasto(desc_lower):
    """
    Clasifica un gasto por descripción.
    Orden de búsqueda:
      1. proveedores.xlsx  (keywords exactas, comparación en MAYÚSCULAS)
      2. GASTO_KEYWORDS    (palabras clave genéricas)
      3. 'Otros'           (fallback)
    """
    desc_upper = desc_lower.upper()
    for keyword, cat in _PROV_MAP.items():
        if keyword in desc_upper:
            return cat
    for cat, kws in GASTO_KEYWORDS:
        if any(k in desc_lower for k in kws):
            return cat
    return 'Otros'

def _normalize_forma(text):
    """Normaliza el texto de una forma de pago a una categoría estándar."""
    t = text.upper()
    if 'EFECTIVO' in t:
        return 'Efectivo'
    if 'YAPPY' in t:
        return 'Yappy'
    if 'ACH' in t or 'TRANSFER' in t:
        return 'ACH / Transferencia'
    if ('TARJETA' in t or re.search(r'\bTC\b', t)
            or 'CREDIT' in t or 'DEBIT' in t):
        return 'Tarjeta'
    return 'Otros'

def _merge_gastos(bg_gastos, gb_gastos):
    """Suma los gastos de Banco General y Global Bank por categoría."""
    merged = {}
    for c in set(list(bg_gastos.keys()) + list(gb_gastos.keys())):
        merged[c] = round(bg_gastos.get(c, 0) + gb_gastos.get(c, 0), 2)
    return merged

# ════════════════════════════════════════════════════════════════════
# HELPERS DE LECTURA DE ARCHIVOS
# ════════════════════════════════════════════════════════════════════

def _to_bytes(path_or_file):
    """
    Acepta:
      • ruta de archivo (str / Path)  → abre y lee
      • file-like object (BytesIO, UploadedFile de Streamlit) → lee bytes
    Retorna bytes para poder crear múltiples BytesIO independientes
    (necesario cuando hay que leer el mismo archivo dos veces, ej. Global Bank).
    """
    if isinstance(path_or_file, (str, os.PathLike)):
        with open(path_or_file, 'rb') as f:
            return f.read()
    data = path_or_file.read()
    if hasattr(path_or_file, 'seek'):
        path_or_file.seek(0)   # deja el stream listo para otro uso
    return data

def _float(s):
    """Convierte '$1,234.56'  o  '1234.56'  a float."""
    try:
        return float(str(s).replace('$', '').replace(',', '').strip())
    except Exception:
        return 0.0

# ════════════════════════════════════════════════════════════════════
# PARSERS
# ════════════════════════════════════════════════════════════════════

# ── Facturas PDF ─────────────────────────────────────────────────
def parse_facturas(path, log=print, paid_by_recibo=None):
    """
    Acepta un solo PDF o una lista de PDFs (ej. una factura por archivo).
    paid_by_recibo: set de int — N. Internos de facturas que ya tienen un
    recibo del mismo día; estas se excluyen de ventas_contado para evitar
    duplicar el ingreso (el recibo las documenta con más detalle).
    Retorna dict con:
      contado_subtotal, contado_itbms,
      credito_subtotal, credito_itbms,
      notas_credito: [{nc, factura, cliente, motivo, subtotal, itbms, total}]
    """
    # Normalizar: siempre trabajar con lista
    paths = path if isinstance(path, list) else [path]

    facturas   = []
    notas_cred = []

    for p in paths:
        data = _to_bytes(p)
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''

                # Fix 1: el PDF dice "NOTA DE CRÉDITO" (con "DE"), no "NOTA CRÉDITO"
                is_nc = bool(re.search(r'NOTA\s+(?:DE\s+)?CR[EÉ]DITO', text, re.IGNORECASE))

                num_m   = re.search(r'N[uú]mero:\s*0*(\d+)',      text)
                sub_all = re.findall(r'Subtotal:\s*([\d,]+\.?\d*)', text)
                itb_m   = re.search(r'ITBMS:\s*([\d,]+\.?\d*)',    text)
                tot_m   = re.search(r'^Total:\s*([\d,]+\.?\d*)',    text, re.MULTILINE)

                numero   = num_m.group(1)          if num_m   else '?'
                # Si hay dos "Subtotal:" (con descuento), tomar el último (ya descontado)
                subtotal = _float(sub_all[-1])     if sub_all else 0.0
                itbms    = _float(itb_m.group(1)) if itb_m  else 0.0
                total    = _float(tot_m.group(1)) if tot_m  else subtotal + itbms

                if is_nc:
                    # Fix 2: el número de NC aparece como "13-000000133" en su propia
                    # línea, sin etiqueta "Número:" → buscarlo como fallback
                    if numero == '?':
                        nc_id_m = re.search(r'^\s*(\d{2}-\d+)\s*$', text, re.MULTILINE)
                        if nc_id_m:
                            numero = nc_id_m.group(1)

                    # Fix 3: "Aplicado A: 11-000001862" es referencia directa (no CUFE)
                    fact_ref = '?'
                    app_m = re.search(r'Aplicado\s+A:\s*(\S+)', text, re.IGNORECASE)
                    if app_m:
                        ref_raw = app_m.group(1)
                        # Formato directo "XX-0001862" → extraer número sin ceros iniciales
                        direct_m = re.search(r'\d+-0*(\d+)', ref_raw)
                        if direct_m:
                            fact_ref = direct_m.group(1)
                        else:
                            # Fallback: formato CUFE largo (facturas antiguas)
                            ref_m = re.search(r'20\d{6}(\d{10})001', ref_raw)
                            if ref_m:
                                fact_ref = str(int(ref_m.group(1)))

                    com_m  = re.search(r'Comentario:\s*(.+)', text)
                    motivo = com_m.group(1).strip() if com_m else ''

                    cli_m   = re.search(r'DV\s+77\s+(.+)', text)
                    cliente = cli_m.group(1).strip() if cli_m else ''

                    notas_cred.append({
                        'nc':       numero,
                        'factura':  fact_ref,
                        'cliente':  cliente,
                        'motivo':   motivo,
                        'subtotal': subtotal,
                        'itbms':    itbms,
                        'total':    total,
                    })
                    log(f"  NC {numero} → Factura {fact_ref} — {cliente}")
                else:
                    # Leer el tipo de pago exclusivamente desde la sección
                    # "FORMAS DE PAGO" — ignora el nombre del cliente ("Contado")
                    # y el campo "Consumidor Final" que pueden inducir errores.
                    pago_m = re.search(
                        r'FORMAS\s+DE\s+PAGO[:\s]+(.*)',
                        text,
                        re.IGNORECASE | re.DOTALL,
                    )
                    if pago_m:
                        pago_text = pago_m.group(1)
                        if re.search(r'\bCR[EÉ]DITO\b', pago_text, re.IGNORECASE):
                            tipo = 'CREDITO'
                        elif re.search(r'\bACH\b|\bTRANSFER', pago_text, re.IGNORECASE):
                            tipo = 'CONTADO_ACH'
                        elif re.search(r'\bEFECTIVO\b', pago_text, re.IGNORECASE):
                            tipo = 'CONTADO_EFECTIVO'
                        else:
                            # Yappy, Tarjeta, Cheque, etc. — el dinero llega
                            # días después, se trata como crédito hasta recibir
                            # el pago (recibo con forma de pago correspondiente).
                            tipo = 'CREDITO'
                    else:
                        # Fallback si la factura no tiene sección FORMAS DE PAGO
                        is_contado = bool(re.search(
                            r'Consumidor\s+Final', text, re.IGNORECASE
                        ))
                        tipo = 'CONTADO_EFECTIVO' if is_contado else 'CREDITO'
                    facturas.append({
                        'numero':   numero,
                        'tipo':     tipo,
                        'subtotal': subtotal,
                        'itbms':    itbms,
                        'total':    total,
                    })
                    # Indicar en el log si la factura quedará en $0.00
                    _n = None
                    try:
                        _n = int(numero)
                    except (ValueError, TypeError):
                        pass
                    if _n is not None and paid_by_recibo and _n in paid_by_recibo:
                        log(f"  Factura {numero} [{tipo}] $0.00 — cubierta por recibo del mismo día")
                    else:
                        log(f"  Factura {numero} [{tipo}] Sub={subtotal:.2f}  ITBMS={itbms:.2f}")

    _paid = paid_by_recibo or set()

    def _excluir(f):
        """True si la factura está cubierta por un recibo del mismo día.
        Las NCs NO excluyen la factura referenciada — se restan directamente
        del total de crédito para cubrir tanto NCs del mismo día como de días
        anteriores (y también NCs parciales)."""
        try:
            return int(f['numero']) in _paid
        except (ValueError, TypeError):
            return False

    # Contado: NUNCA excluir — la venta ocurrió en el momento.
    # El recibo adjunto a una factura contado es solo un comprobante de pago,
    # no un cobro de cartera; aplicar _excluir aquí causaría que ventas
    # contado-ACH o contado-efectivo desaparezcan del informe.
    contado_ef_sub  = sum(f['subtotal'] for f in facturas
                          if f['tipo'] == 'CONTADO_EFECTIVO')
    contado_ef_itb  = sum(f['itbms']    for f in facturas
                          if f['tipo'] == 'CONTADO_EFECTIVO')
    contado_ach_sub = sum(f['subtotal'] for f in facturas
                          if f['tipo'] == 'CONTADO_ACH')
    contado_ach_itb = sum(f['itbms']    for f in facturas
                          if f['tipo'] == 'CONTADO_ACH')
    # Crédito: sí excluir si el cliente pagó el mismo día (ya aparece en Cobros).
    credito_sub     = sum(f['subtotal'] for f in facturas
                          if f['tipo'] == 'CREDITO' and not _excluir(f))
    credito_itb     = sum(f['itbms']    for f in facturas
                          if f['tipo'] == 'CREDITO' and not _excluir(f))

    # Restar NCs directamente — funciona para NCs del mismo día Y de días
    # anteriores, y también para anulaciones parciales de una factura.
    credito_sub = round(credito_sub - sum(nc['subtotal'] for nc in notas_cred), 2)
    credito_itb = round(credito_itb - sum(nc['itbms']    for nc in notas_cred), 2)

    return {
        'contado_efectivo_subtotal': round(contado_ef_sub,  2),
        'contado_efectivo_itbms':    round(contado_ef_itb,  2),
        'contado_ach_subtotal':      round(contado_ach_sub, 2),
        'contado_ach_itbms':         round(contado_ach_itb, 2),
        # Agregados para compatibilidad con código legado
        'contado_subtotal': round(contado_ef_sub + contado_ach_sub, 2),
        'contado_itbms':    round(contado_ef_itb + contado_ach_itb, 2),
        'credito_subtotal': round(credito_sub, 2),
        'credito_itbms':    round(credito_itb, 2),
        'notas_credito':    notas_cred,
    }

# ── Helpers recibos ──────────────────────────────────────────────
def _get_paid_factura_refs(path, log=print):
    """
    Lee los PDFs de recibos y retorna un set de int con los N. Internos
    de las facturas que ya fueron cobradas hoy (referenciadas en APLICADO A).
    Usado por parse_facturas para excluir esas facturas de ventas_contado;
    el recibo documenta el ingreso con la forma de pago detallada.
    """
    if path is None:
        return set()
    paths = path if isinstance(path, list) else [path]
    refs = set()
    for p in (paths if paths else []):
        if p is None:
            continue
        data = _to_bytes(p)
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        if re.search(r'APLICADO\s+A:.*?Factura', text, re.IGNORECASE | re.DOTALL):
            num = _extract_factura_num_from_recibo(text)
            if num is not None:
                refs.add(num)
                log(f"  [pre-scan] Recibo cubre factura #{num} — se excluirá de ventas contado")
    return refs

def _extract_factura_num_from_recibo(text):
    """
    Extrae como int el número de la factura referenciada en APLICADO A.
    En el PDF, el N. Interno queda partido en dos líneas:
        "11-  FE01...  13-07-"
        "000001858  77000126...  2026"
    Buscamos la segunda parte: un número de ≥4 dígitos seguido del
    fragmento largo del CUFE (≥20 dígitos).
    Devuelve int o None si no se encuentra.
    """
    blk_m = re.search(
        r'APLICADO\s+A:(.*?)(?:FORMAS\s+DE\s+PAGO|OBSERVACIONES|$)',
        text, re.IGNORECASE | re.DOTALL
    )
    if not blk_m:
        return None
    for line in blk_m.group(1).split('\n'):
        m = re.match(r'^(\d{4,})\s+\d{20,}', line.strip())
        if m:
            return int(m.group(1))
    return None

# ── Recibos PDF ──────────────────────────────────────────────────
def parse_recibos(path, log=print):
    """
    Acepta un solo PDF o una lista de PDFs.
    Retorna dict con:
      'total'    : float — suma de todos los cobros a clientes
      'por_forma': dict  — {forma: monto} para cada forma en FORMAS_ORDEN

    Todos los recibos se cuentan: A Cuenta, efectivo, Yappy, TC, ACH…
    La deduplicación se hace en parse_facturas: las facturas que ya tienen
    un recibo del mismo día se excluyen de ventas_contado (via paid_by_recibo).

    Tipos reconocidos:
    - A Cuenta: extrae monto de la fila "A Cuenta DD-MM-YYYY monto"
    - Aplicado a factura: busca "Totales:" o suma montos con saldo 0.00
    """
    paths = path if isinstance(path, list) else [path]
    total    = 0.0
    por_forma = {f: 0.0 for f in FORMAS_ORDEN}

    for p in paths:
        data = _to_bytes(p)
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

        subtotal = 0.0

        # ── Caso 1: Recibo "A Cuenta" (prepago) ─────────────────────────
        acuenta_m = re.search(
            r'A\s+Cuenta\s+\d{2}[/-]\d{2}[/-]\d{4}\s+([\d,]+\.\d{2})',
            text, re.IGNORECASE
        )
        if acuenta_m:
            subtotal = _float(acuenta_m.group(1))
            log(f"  Recibo A Cuenta: ${subtotal:,.2f}")

        else:
            # ── Caso 2: Recibo aplicado a factura (cualquier forma de pago) ─
            tot_m = re.search(r'Totales?:\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
            if tot_m:
                subtotal = _float(tot_m.group(1))
                log(f"  Recibo cobro: ${subtotal:,.2f}")
            else:
                items = re.findall(r'\b(\d{1,3}(?:,\d{3})*\.\d{2})\s+0\.00', text)
                subtotal = sum(_float(t) for t in items)
                log(f"  Recibo cobro (suma individual): ${subtotal:,.2f}")

        total += subtotal

        # ── Extraer forma de pago del bloque FORMAS DE PAGO ─────────────
        formas_blk = re.search(
            r'FORMAS\s+DE\s+PAGO[:\s]+(.*?)(?:OBSERVACIONES|$)',
            text, re.IGNORECASE | re.DOTALL
        )
        if formas_blk:
            for line in formas_blk.group(1).split('\n'):
                line = line.strip()
                if not line:
                    continue
                amt_m = re.search(r'([\d,]+\.\d{2})', line)
                if not amt_m:
                    continue
                prefix = line[:amt_m.start()].strip()
                # Saltar líneas de encabezado de la tabla
                if re.search(r'Forma\s+de\s+Pago|Concepto|Descripci', prefix, re.IGNORECASE):
                    continue
                amt = _float(amt_m.group(1))
                if amt > 0 and prefix:
                    forma = _normalize_forma(prefix)
                    por_forma[forma] = round(por_forma[forma] + amt, 2)
                    log(f"  Forma de pago: {forma} ${amt:.2f} (de: {prefix[:40]})")

    log(f"  Total cobros acumulado: ${total:,.2f}")
    return {
        'total':     round(total, 2),
        'por_forma': por_forma,
    }

# ── Banco General PDF ────────────────────────────────────────────
def parse_banco_general(path, log=print):
    """
    Retorna:
      saldo_anterior, depositos,
      gastos: {categoria: monto},
      gastos_detalle: [{desc, monto, cat}]
    """
    if path is None:
        log("  Sin extracto Banco General — se omite.")
        return {'saldo_anterior': 0, 'depositos': 0,
                'gastos': {c: 0.0 for c in ITEMS_PAGOS}, 'gastos_detalle': []}
    data = _to_bytes(path)
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

    pairs = re.findall(r'(-?\$[\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})', text)
    if not pairs:
        log("  ADVERTENCIA: no se encontraron transacciones en extracto BG")
        return {'saldo_anterior': 0, 'depositos': 0,
                'gastos': {c: 0.0 for c in ITEMS_PAGOS}, 'gastos_detalle': []}

    amounts = [_float(p[0]) for p in pairs]
    saldos  = [_float(p[1]) for p in pairs]

    saldo_anterior = round(saldos[0] - amounts[0], 2)
    depositos      = round(sum(a for a in amounts if a > 0), 2)

    gastos_por_cat = {c: 0.0 for c in ITEMS_PAGOS}
    gastos_detalle = []
    neg_pattern = re.compile(r'(-\$[\d,]+\.\d{2})\s+\$[\d,]+\.\d{2}')

    for line in text.split('\n'):
        m = neg_pattern.search(line)
        if m:
            monto = abs(_float(m.group(1)))
            desc  = line[:m.start()].strip()
            cat   = _classify_gasto(desc.lower())
            gastos_por_cat[cat] = round(gastos_por_cat.get(cat, 0) + monto, 2)
            gastos_detalle.append({'desc': desc, 'monto': monto, 'cat': cat})
            log(f"  BG gasto [{cat}]: ${monto:.2f} — {desc[:60]}")

    log(f"  BG saldo anterior: ${saldo_anterior:,.2f}  depósitos: ${depositos:,.2f}")
    return {
        'saldo_anterior': saldo_anterior,
        'depositos':      depositos,
        'gastos':         gastos_por_cat,
        'gastos_detalle': gastos_detalle,
    }

# ── Global Bank XLS ──────────────────────────────────────────────
def parse_global_bank(path, log=print):
    """
    Retorna:
      saldo_anterior, depositos,
      gastos: {categoria: monto},
      gastos_detalle: [{desc, monto, cat}]
    """
    if path is None:
        log("  Sin extracto Global Bank — se omite.")
        return {'saldo_anterior': 0, 'depositos': 0,
                'gastos': {c: 0.0 for c in ITEMS_PAGOS}, 'gastos_detalle': []}
    # Leer bytes una vez, crear dos BytesIO independientes para las dos pasadas
    raw = _to_bytes(path)

    df = pd.read_excel(io.BytesIO(raw), engine='xlrd', header=None)

    # Encontrar la fila de encabezado (contiene "Débitos" / "Créditos")
    hdr_row = None
    for i, row in df.iterrows():
        vals = [str(v).strip() for v in row.values]
        if any('bito' in v.lower() for v in vals):
            hdr_row = i
            break

    if hdr_row is None:
        log("  ADVERTENCIA: no se encontró encabezado en Global Bank XLS")
        return {'saldo_anterior': 0, 'depositos': 0,
                'gastos': {c: 0.0 for c in ITEMS_PAGOS}, 'gastos_detalle': []}

    # Segunda pasada con encabezado correcto
    df2 = pd.read_excel(io.BytesIO(raw), engine='xlrd', header=hdr_row)
    df2.columns = [str(c).strip() for c in df2.columns]

    def find_col(keywords):
        for c in df2.columns:
            if any(k in c.lower() for k in keywords):
                return c
        return None

    col_debito  = find_col(['bito', 'débito', 'cargo'])
    col_credito = find_col(['crédito', 'credito', 'abono'])
    col_saldo   = find_col(['saldo'])
    col_desc    = find_col(['concepto', 'descripci', 'detalle'])

    if not all([col_debito, col_credito, col_saldo]):
        log("  ADVERTENCIA: columnas no encontradas en Global Bank XLS")
        return {'saldo_anterior': 0, 'depositos': 0,
                'gastos': {c: 0.0 for c in ITEMS_PAGOS}, 'gastos_detalle': []}

    data = df2.dropna(subset=[col_saldo]).copy()
    data = data[pd.to_numeric(data[col_saldo], errors='coerce').notna()]

    if data.empty:
        return {'saldo_anterior': 0, 'depositos': 0,
                'gastos': {c: 0.0 for c in ITEMS_PAGOS}, 'gastos_detalle': []}

    debitos  = pd.to_numeric(data[col_debito],  errors='coerce').fillna(0)
    creditos = pd.to_numeric(data[col_credito], errors='coerce').fillna(0)
    saldos   = pd.to_numeric(data[col_saldo],   errors='coerce').fillna(0)

    total_debitos  = round(float(debitos.sum()),  2)
    total_creditos = round(float(creditos.sum()), 2)

    # Detectar orden cronológico vs. más-reciente-primero con identidad contable
    def _sfloat(v):
        f = pd.to_numeric(v, errors='coerce')
        return 0.0 if (f != f) else float(f)

    data_r = data.reset_index(drop=True)
    oldest_first_score = newest_first_score = 0
    for i in range(len(data_r) - 1):
        s0  = _sfloat(data_r.iloc[i][col_saldo])
        s1  = _sfloat(data_r.iloc[i+1][col_saldo])
        cr1 = _sfloat(data_r.iloc[i+1][col_credito])
        db1 = _sfloat(data_r.iloc[i+1][col_debito])
        cr0 = _sfloat(data_r.iloc[i][col_credito])
        db0 = _sfloat(data_r.iloc[i][col_debito])
        if abs(s1 - s0 - cr1 + db1) < 0.02:
            oldest_first_score += 1
        if abs(s0 - s1 - cr0 + db0) < 0.02:
            newest_first_score += 1

    if oldest_first_score >= newest_first_score:
        final_saldo = float(list(saldos)[-1])
    else:
        final_saldo = float(list(saldos)[0])

    saldo_ant = round(final_saldo - total_creditos + total_debitos, 2)
    depositos = total_creditos

    gastos_por_cat = {c: 0.0 for c in ITEMS_PAGOS}
    gastos_detalle = []

    for _, row in data.iterrows():
        deb = float(pd.to_numeric(row[col_debito], errors='coerce') or 0)
        if deb > 0:
            raw_desc = row.get(col_desc, '') if col_desc else ''
            desc     = '' if (raw_desc != raw_desc) else str(raw_desc)
            cat      = _classify_gasto(desc.lower())
            gastos_por_cat[cat] = round(gastos_por_cat.get(cat, 0) + deb, 2)
            gastos_detalle.append({'desc': desc, 'monto': deb, 'cat': cat})
            log(f"  GB gasto [{cat}]: ${deb:.2f} — {desc[:60]}")

    log(f"  GB saldo anterior: ${saldo_ant:,.2f}  depósitos: ${depositos:,.2f}")
    return {
        'saldo_anterior': saldo_ant,
        'depositos':      depositos,
        'gastos':         gastos_por_cat,
        'gastos_detalle': gastos_detalle,
    }

# ════════════════════════════════════════════════════════════════════
# GENERADOR DEL REPORTE EXCEL
# ════════════════════════════════════════════════════════════════════

def generate_report(fecha_str, facturas, recibos, bg, gb,
                    output_path=None, log=print,
                    cheques_bg=0.0, cheques_gb=0.0,
                    cobros_por_forma=None,
                    logo_path=None):
    """
    Genera el Cierre Diario de Caja (Excel).

    Si output_path es None  → retorna un BytesIO listo para descargar (Streamlit).
    Si output_path es str   → guarda en disco y retorna None              (tkinter).

    cobros_por_forma : dict {forma: monto} con las 5 claves de FORMAS_ORDEN.
    logo_path        : ruta al logo (JPG/PNG); si None, se usa texto como fallback.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cierre Diario"

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13

    def spacer(row, h=8):
        ws.row_dimensions[row].height = h

    def section_hdr(row, title):
        ws.merge_cells(f'A{row}:D{row}')
        c = ws[f'A{row}']
        c.value     = title
        c.font      = Font(name='Garamond', bold=True, size=11, color=WHITE)
        c.fill      = mk_fill(DARK_BLUE)
        c.alignment = CA
        ws.row_dimensions[row].height = 22

    def col_hdr(row, headers):
        for col, hdr in zip(['A', 'B', 'C', 'D'], headers):
            c = ws[f'{col}{row}']
            c.value     = hdr
            c.font      = Font(name='Garamond', bold=True, size=11)
            c.fill      = mk_fill(LIGHT_BLUE)
            c.alignment = LA if col == 'A' else CA
            c.border    = BRD
        ws.row_dimensions[row].height = 20

    def row_s1(row, label, b=None, c=None, d=None, bold=False, bg_color=None):
        ws.row_dimensions[row].height = 20
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.border    = BRD
            cell.font      = Font(name='Garamond', bold=bold, size=11)
            cell.alignment = LA if col == 'A' else RA
            if col != 'A':
                cell.number_format = CURRENCY
            if bg_color:
                cell.fill = mk_fill(bg_color)
        ws[f'A{row}'].value = label
        if b is not None: ws[f'B{row}'].value = b
        if c is not None: ws[f'C{row}'].value = c
        if d is not None: ws[f'D{row}'].value = d

    def row_s2(row, label, d=None, bold=False, bg_color=None):
        ws.row_dimensions[row].height = 20
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.border    = BRD
            cell.font      = Font(name='Garamond', bold=bold, size=11)
            cell.alignment = LA if col == 'A' else RA
            if col == 'D':
                cell.number_format = CURRENCY
            if bg_color:
                cell.fill = mk_fill(bg_color)
        ws[f'A{row}'].value = label
        if d is not None: ws[f'D{row}'].value = d

    def row_s3(row, label, b=None, c=None, bold=False, bg_color=None):
        ws.row_dimensions[row].height = 20
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.border    = BRD
            cell.font      = Font(name='Garamond', bold=bold, size=11)
            cell.alignment = LA if col == 'A' else RA
            if col in ['B', 'C']:
                cell.number_format = CURRENCY
            if bg_color:
                cell.fill = mk_fill(bg_color)
        ws[f'A{row}'].value = label
        if b is not None: ws[f'B{row}'].value = b
        if c is not None: ws[f'C{row}'].value = c

    # ── Encabezado — fondo azul oscuro + logo ──────────────────────
    # Fila 1: logo (fila alta, relleno azul oscuro)
    ws.row_dimensions[1].height = 72
    ws.merge_cells('A1:D1')
    ws['A1'].fill      = mk_fill(DARK_BLUE)
    ws['A1'].alignment = CA

    if logo_path and os.path.exists(logo_path) and _HAS_XL_IMAGE:
        try:
            img = XLImage(logo_path)
            # Original 1122×450 px → mostrar ~250×100 px (proporción ~2.49:1)
            img.width  = 250
            img.height = 100
            img.anchor = 'A1'
            ws.add_image(img)
        except Exception:
            ws['A1'].value = 'GRUPO PANABLOCK'
            ws['A1'].font  = Font(name='Garamond', bold=True, size=14, color=WHITE)
    else:
        ws['A1'].value = 'GRUPO PANABLOCK'
        ws['A1'].font  = Font(name='Garamond', bold=True, size=16, color=WHITE)

    # Fila 2: título en blanco sobre fondo azul
    ws.row_dimensions[2].height = 26
    ws.merge_cells('A2:D2')
    ws['A2'].value     = 'Informe de Cierre Diario de Caja'
    ws['A2'].font      = Font(name='Garamond', bold=True, size=18, color=WHITE)
    ws['A2'].fill      = mk_fill(DARK_BLUE)
    ws['A2'].alignment = CA

    # Fila 3: fecha en blanco sobre fondo azul
    ws.row_dimensions[3].height = 20
    ws.merge_cells('A3:D3')
    ws['A3'].value     = f'Fecha: {fecha_str}'
    ws['A3'].font      = Font(name='Garamond', italic=True, size=11, color=WHITE)
    ws['A3'].fill      = mk_fill(DARK_BLUE)
    ws['A3'].alignment = LA

    spacer(4)

    # ── Sección 1 — VENTAS Y COBROS ────────────────────────────────
    section_hdr(5, 'VENTAS Y COBROS')
    spacer(6, 8)

    # ── Pre-cómputo de todos los totales ───────────────────────────
    _ef_sub  = facturas['contado_efectivo_subtotal']
    _ef_itb  = facturas['contado_efectivo_itbms']
    _ach_sub = facturas['contado_ach_subtotal']
    _ach_itb = facturas['contado_ach_itbms']
    _cr_sub  = facturas['credito_subtotal']
    _cr_itb  = facturas['credito_itbms']
    _ef_tot  = round(_ef_sub  + _ef_itb,  2)
    _ach_tot = round(_ach_sub + _ach_itb, 2)
    _cr_tot  = round(_cr_sub  + _cr_itb,  2)
    _b11     = round(_ef_sub  + _ach_sub  + _cr_sub, 2)
    _c11     = round(_ef_itb  + _ach_itb  + _cr_itb, 2)
    _d11     = round(_b11 + _c11, 2)
    _pf      = cobros_por_forma or {f: 0.0 for f in FORMAS_ORDEN}
    _cob_ef  = _pf.get('Efectivo', 0.0)
    _d19     = round(sum(_pf.values()), 2)
    _d21     = round(_ef_tot + _cob_ef, 2)

    # Fila 7: encabezado columnas ventas (A7 bold size=14)
    ws.row_dimensions[7].height = 20
    for col, hdr in zip(['A', 'B', 'C', 'D'], ['Ventas', 'Subtotal', 'ITBMS', 'Total']):
        c = ws[f'{col}7']
        c.value     = hdr
        c.font      = Font(name='Garamond', bold=True, size=14 if col == 'A' else 11)
        c.fill      = mk_fill(LIGHT_BLUE)
        c.alignment = LA if col == 'A' else CA
        c.border    = BRD

    # Filas 8-10: ventas por categoría (indent=2 en columna A)
    row_s1(8,  'Ventas al Contado - EFECTIVO',
           b=_ef_sub,  c=_ef_itb,  d=_ef_tot  if _ef_tot  else None)
    ws['A8'].alignment = LA2
    row_s1(9,  'Ventas al Contado - ACH',
           b=_ach_sub, c=_ach_itb, d=_ach_tot if _ach_tot else None)
    ws['A9'].alignment = LA2
    row_s1(10, 'Ventas al Crédito',
           b=_cr_sub,  c=_cr_itb,  d=_cr_tot  if _cr_tot  else None)
    ws['A10'].alignment = LA2

    # Fila 11: Total Ventas del día
    row_s1(11, 'Total Ventas del día',
           b=_b11 if _b11 else None,
           c=_c11 if _c11 else None,
           d=_d11 if _d11 else None,
           bold=True, bg_color=LGRAY)

    spacer(12, 10)

    # Fila 13: encabezado sección Cobros
    ws.row_dimensions[13].height = 18
    for col, hdr in zip(['A', 'B', 'C', 'D'], ['Cobros', '', '', '']):
        c = ws[f'{col}13']
        c.value     = hdr
        c.font      = Font(name='Garamond', bold=True, size=14 if col == 'A' else 11)
        c.fill      = mk_fill(LIGHT_BLUE)
        c.alignment = LA if col == 'A' else CA
        c.border    = BRD

    # Filas 14-18: cobros por forma de pago
    for i, forma in enumerate(FORMAS_ORDEN):
        r = 14 + i
        ws.row_dimensions[r].height = 18
        amt = _pf.get(forma, 0.0)
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{r}']
            cell.border    = BRD
            cell.font      = Font(name='Garamond', size=11)
            cell.alignment = LA2 if col == 'A' else RA
            if col == 'D':
                cell.number_format = CURRENCY
        ws[f'A{r}'].value = forma
        ws[f'D{r}'].value = amt if amt else None

    # Fila 19: Total Cobrado en el día
    row_s1(19, 'Total Cobrado en el día',
           d=_d19 if _d19 else None, bold=True, bg_color=LGRAY)
    ws.row_dimensions[19].height = 18

    spacer(20, 8)

    # Fila 21: Total a Depositar en el Día = Ventas Efectivo + Cobros Efectivo
    ws.row_dimensions[21].height = 20
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}21']
        cell.border    = BRD
        cell.fill      = mk_fill(MGRAY)
        cell.alignment = LA if col == 'A' else RA
        if col == 'D':
            cell.number_format = CURRENCY
            cell.font = Font(name='Garamond', bold=True, size=12)
        else:
            cell.font = Font(name='Garamond', bold=True, size=11)
    ws['A21'].value = 'Total a Depositar en el Día'
    ws['D21'].value = _d21 if _d21 else None

    spacer(22)
    spacer(23, 8)

    # ── Sección 2 — PAGOS Y GASTOS ─────────────────────────────────
    section_hdr(24, 'PAGOS Y GASTOS')
    col_hdr(25, ['Descripción', 'Banco General', 'Global Bank', 'Total'])

    gastos     = _merge_gastos(bg['gastos'], gb['gastos'])
    total_gast = round(sum(gastos.values()), 2)
    bg_egresos = round(sum(bg['gastos'].values()), 2)
    gb_egresos = round(sum(gb['gastos'].values()), 2)
    row_s1(26, 'Pago a Proveedores',
           b=bg_egresos if bg_egresos else None,
           c=gb_egresos if gb_egresos else None,
           d=total_gast if total_gast else None,
           bold=True, bg_color=LGRAY)

    spacer(27)

    # ── Sección 3 — MOVIMIENTOS DE BANCOS ──────────────────────────
    section_hdr(28, 'MOVIMIENTOS DE BANCOS')
    col_hdr(29, ['Descripción', 'Banco General', 'Global Bank', ''])

    row_s3(30, 'Saldo Anterior',
           b=bg['saldo_anterior'], c=gb['saldo_anterior'])
    row_s3(31, 'Depósitos según extracto bancario (efectivo, ACH, TC, Yappy)',
           b=bg['depositos'], c=gb['depositos'])
    ws.row_dimensions[31].height = 36

    row_s3(32, 'Egresos según extracto bancario',
           b=bg_egresos if bg_egresos else None,
           c=gb_egresos if gb_egresos else None)

    _bg_sf = round(bg['saldo_anterior'] + bg['depositos'] - bg_egresos, 2)
    _gb_sf = round(gb['saldo_anterior'] + gb['depositos'] - gb_egresos, 2)
    _bg_rc = round(_bg_sf - (cheques_bg or 0), 2)
    _gb_rc = round(_gb_sf - (cheques_gb or 0), 2)

    row_s3(33, 'Saldo Bancario al Final del Día',
           b=_bg_sf if _bg_sf else None,
           c=_gb_sf if _gb_sf else None,
           bold=True, bg_color=LGRAY)

    row_s3(34, 'Cheques en Circulación',
           b=cheques_bg if cheques_bg else None,
           c=cheques_gb if cheques_gb else None)

    row_s3(35, 'Saldo Bancario Real Conciliado',
           b=_bg_rc if _bg_rc else None,
           c=_gb_rc if _gb_rc else None,
           bold=True, bg_color=MGRAY)
    ws['D34'].border = NO_BORDER

    spacer(36)

    # ── Firma ───────────────────────────────────────────────────────
    ws.row_dimensions[37].height = 30
    ws.merge_cells('A37:D37')
    ws['A37'].value     = 'Firma: ___________________________'
    ws['A37'].font      = Font(name='Garamond', size=11)
    ws['A37'].alignment = LA

    spacer(38)

    # ── Leyenda ─────────────────────────────────────────────────────
    ws.row_dimensions[39].height = 16
    ws.merge_cells('A39:D39')
    ws['A39'].value     = ('Generado automáticamente por Cierre de Caja — Grupo Panablock')
    ws['A39'].font      = Font(name='Garamond', italic=True, size=9, color='595959')
    ws['A39'].alignment = LA

    ws.freeze_panes = 'A8'

    # ── Guardar ─────────────────────────────────────────────────────
    if output_path is None:
        # Modo Streamlit: retornar BytesIO para descarga directa
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    else:
        # Modo escritorio: guardar en disco
        wb.save(output_path)
        log(f"\n✔ Reporte guardado en: {output_path}")
        return None


# ════════════════════════════════════════════════════════════════════
# GENERADOR DEL REPORTE PDF
# ════════════════════════════════════════════════════════════════════

def generate_pdf_report(fecha_str, facturas, recibos, bg, gb,
                        output_path=None, log=print,
                        cheques_bg=0.0, cheques_gb=0.0,
                        cobros_por_forma=None,
                        logo_path=None):
    """
    Genera el Cierre Diario de Caja en formato PDF (ReportLab/Platypus).
    Orientación portrait, tamaño carta.

    Si output_path es None → retorna BytesIO para descarga (Streamlit).
    Si output_path es str  → guarda en disco y retorna None.
    """
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, Image,
    )
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    # ── Colores ────────────────────────────────────────────────────
    C_DARK   = HexColor('#1F4E79')   # azul oscuro — encabezados de sección
    C_MED    = HexColor('#2E75B6')   # azul medio  — encabezados de columna
    C_PALE   = HexColor('#D9E8F5')   # azul pálido — sub-encabezado desglose
    C_VLIGHT = HexColor('#EBF1F8')   # azul muy claro — filas desglose
    C_LGRAY  = HexColor('#D9D9D9')   # gris claro  — filas de subtotal
    C_MGRAY  = HexColor('#BFBFBF')   # gris medio  — filas de total principal
    C_RED    = HexColor('#FF0000')   # rojo — sección NC
    C_DKRED  = HexColor('#CC0000')   # rojo oscuro — encabezado col NC
    C_REDP   = HexColor('#FFE0E0')   # rosa — filas NC
    C_WHITE  = colors.white
    C_GRAY6  = HexColor('#595959')

    # ── Layout ────────────────────────────────────────────────────
    PAGE_W, PAGE_H = letter          # 612 × 792 pt
    MARGIN = 0.50 * inch             # 36 pt — márgenes reducidos para caber en 1 página
    CW = PAGE_W - 2 * MARGIN         # 540 pt

    FONT   = 'Helvetica'
    FONT_B = 'Helvetica-Bold'
    FONT_I = 'Helvetica-Oblique'

    # ── Helpers numéricos ─────────────────────────────────────────
    def money(v):
        """'—' para None/0; '$X,XXX.XX' o '-$X,XXX.XX' para cualquier otro."""
        if not v:
            return '—'
        return f'${v:,.2f}' if v > 0 else f'-${abs(v):,.2f}'

    def money_show(v):
        """Siempre muestra el valor, incluyendo $0.00."""
        if v is None:
            return '$0.00'
        if v < 0:
            return f'-${abs(v):,.2f}'
        return f'${v:,.2f}'

    # ── Pre-cálculos ──────────────────────────────────────────────
    pf = cobros_por_forma or {f: 0.0 for f in FORMAS_ORDEN}

    ef_sub   = facturas['contado_efectivo_subtotal']
    ef_itb   = facturas['contado_efectivo_itbms']
    ach_sub  = facturas['contado_ach_subtotal']
    ach_itb  = facturas['contado_ach_itbms']
    credito_sub = facturas['credito_subtotal']
    credito_itb = facturas['credito_itbms']

    ef_total      = round(ef_sub  + ef_itb,  2)
    ach_total     = round(ach_sub + ach_itb, 2)
    total_credito = round(credito_sub + credito_itb, 2)
    sum_sub       = round(ef_sub + ach_sub + credito_sub, 2)
    sum_itb       = round(ef_itb + ach_itb + credito_itb, 2)
    sum_total     = round(sum_sub + sum_itb, 2)

    cobros_ef       = pf.get('Efectivo', 0.0)
    total_cobros    = round(sum(pf.values()), 2)
    total_depositar = round(ef_total + cobros_ef, 2)

    # Compat aliases (used in Sección 2/3 below)
    contado_sub = round(ef_sub + ach_sub, 2)
    contado_itb = round(ef_itb + ach_itb, 2)
    total_contado = round(contado_sub + contado_itb, 2)
    total_itbms = round(ef_itb + ach_itb + credito_itb, 2)

    gastos     = _merge_gastos(bg['gastos'], gb['gastos'])
    total_gast = round(sum(gastos.values()), 2)

    bg_egresos = round(sum(bg['gastos'].values()), 2)
    gb_egresos = round(sum(gb['gastos'].values()), 2)

    bg_saldo_fin  = round(bg['saldo_anterior'] + bg['depositos'] - bg_egresos, 2)
    gb_saldo_fin  = round(gb['saldo_anterior'] + gb['depositos'] - gb_egresos, 2)
    bg_conciliado = round(bg_saldo_fin - cheques_bg, 2)
    gb_conciliado = round(gb_saldo_fin - cheques_gb, 2)

    ncs = facturas.get('notas_credito', [])

    # ── Anchos de columna ─────────────────────────────────────────
    # Sección 1: Descripción | Subtotal | ITBMS | Total
    cw1 = [CW * 0.46, CW * 0.18, CW * 0.18, CW * 0.18]
    # Sección 2: Descripción | Total  (2 columnas)
    cw2 = [CW * 0.72, CW * 0.28]
    # Sección 3: Descripción | Banco General | Global Bank
    cw3 = [CW * 0.50, CW * 0.25, CW * 0.25]

    # ── Estilos base ──────────────────────────────────────────────
    BASE = [
        ('FONTNAME',      (0, 0), (-1, -1), FONT),
        ('FONTSIZE',      (0, 0), (-1, -1), 8),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('ALIGN',         (0, 0), (0, -1),  'LEFT'),
        ('GRID',          (0, 0), (-1, -1), 0.5, HexColor('#C8C8C8')),
        ('BOX',           (0, 0), (-1, -1), 1.0, C_DARK),
    ]

    def sec_hdr(n_cols):
        """Estilo fila 0 = encabezado de sección (span completo, azul oscuro)."""
        return [
            ('SPAN',          (0, 0), (n_cols - 1, 0)),
            ('BACKGROUND',    (0, 0), (n_cols - 1, 0), C_DARK),
            ('TEXTCOLOR',     (0, 0), (n_cols - 1, 0), C_WHITE),
            ('FONTNAME',      (0, 0), (n_cols - 1, 0), FONT_B),
            ('FONTSIZE',      (0, 0), (n_cols - 1, 0), 10),
            ('ALIGN',         (0, 0), (n_cols - 1, 0), 'CENTER'),
            ('TOPPADDING',    (0, 0), (n_cols - 1, 0), 4),
            ('BOTTOMPADDING', (0, 0), (n_cols - 1, 0), 4),
        ]

    def col_hdr(row, n_cols):
        """Estilo fila de encabezados de columna (azul medio)."""
        return [
            ('BACKGROUND',    (0, row), (n_cols - 1, row), C_MED),
            ('TEXTCOLOR',     (0, row), (n_cols - 1, row), C_WHITE),
            ('FONTNAME',      (0, row), (n_cols - 1, row), FONT_B),
            ('FONTSIZE',      (0, row), (n_cols - 1, row), 8),
            ('ALIGN',         (1, row), (n_cols - 1, row), 'RIGHT'),
            ('TOPPADDING',    (0, row), (n_cols - 1, row), 3),
            ('BOTTOMPADDING', (0, row), (n_cols - 1, row), 3),
        ]

    def num_right(col0, col1, row0, row1):
        """Alinea a la derecha un bloque de celdas."""
        return [('ALIGN', (col0, row0), (col1, row1), 'RIGHT')]

    def bold_bg(r0, r1, c0, c1, bg_color):
        """Aplica negrita y color de fondo a un rango."""
        return [
            ('BACKGROUND', (c0, r0), (c1, r1), bg_color),
            ('FONTNAME',   (c0, r0), (c1, r1), FONT_B),
        ]

    # ── STORY ────────────────────────────────────────────────────
    story = []

    # ══════════════════════════════════════════════════════════════
    # ENCABEZADO — logo izquierda, título + fecha derecha
    # ══════════════════════════════════════════════════════════════
    LOGO_W = 2.0 * inch
    LOGO_H = round(LOGO_W * 450 / 1122, 1)   # mantiene proporción 1122:450

    title_ps = ParagraphStyle('hdr_title', fontName=FONT_B, fontSize=12,
                              textColor=C_WHITE, alignment=TA_RIGHT, leading=15)
    date_ps  = ParagraphStyle('hdr_date',  fontName=FONT_I, fontSize=10,
                              textColor=HexColor('#BDD7EE'),
                              alignment=TA_RIGHT, leading=14)

    if logo_path and os.path.exists(logo_path):
        logo_cell = Image(logo_path, width=LOGO_W, height=LOGO_H)
    else:
        logo_cell = Paragraph(
            '<b>GRUPO PANABLOCK</b>',
            ParagraphStyle('lbl_fb', fontName=FONT_B, fontSize=16,
                           textColor=C_WHITE, alignment=TA_LEFT),
        )

    hdr_table = Table(
        [[logo_cell,
          [Paragraph('Informe de Cierre Diario de Caja', title_ps),
           Spacer(1, 4),
           Paragraph(f'Fecha: {fecha_str}', date_ps)]]],
        colWidths=[LOGO_W, CW - LOGO_W],
    )
    hdr_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), C_DARK),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING',   (0, 0), (0, -1),   6),
        ('RIGHTPADDING',  (-1, 0), (-1, -1),  8),
        ('BOX',           (0, 0), (-1, -1),  2, C_DARK),
    ]))
    story.append(hdr_table)
    story.append(Spacer(1, 0.10 * inch))

    # ══════════════════════════════════════════════════════════════
    # SECCIÓN 1 — VENTAS Y COBROS
    # Row 0:   encabezado de sección (span 4)
    # Row 1:   encabezados de columna
    # Row 2:   Ventas al Contado - EFECTIVO
    # Row 3:   Ventas al Contado - ACH
    # Row 4:   Ventas al Crédito
    # Row 5:   Total Ventas del día      LGRAY bold
    # Row 6:   encabezado COBROS (sub-sección, span 4)  PALE
    # Row 7-11: formas de pago (A-C span; D = monto)   VLIGHT
    # Row 12:  Total Cobrado en el día   LGRAY bold
    # Row 13:  Total a Depositar         MGRAY bold
    # ══════════════════════════════════════════════════════════════
    data1 = [
        # 0
        ['VENTAS Y COBROS', '', '', ''],
        # 1
        ['Ventas', 'Subtotal', 'ITBMS', 'Total'],
        # 2
        ['Ventas al Contado - EFECTIVO',
         money(ef_sub), money(ef_itb), money(ef_total)],
        # 3
        ['Ventas al Contado - ACH',
         money(ach_sub), money(ach_itb), money(ach_total)],
        # 4
        ['Ventas al Crédito',
         money(credito_sub), money(credito_itb), money(total_credito)],
        # 5
        ['Total Ventas del día',
         money_show(sum_sub), money_show(sum_itb), money_show(sum_total)],
        # 6 — sub-encabezado Cobros
        ['Cobros', '', '', ''],
        # 7-11 — formas de pago
        ['Efectivo',            '', '', money(pf.get('Efectivo', 0))],
        ['ACH / Transferencia', '', '', money(pf.get('ACH / Transferencia', 0))],
        ['Yappy',               '', '', money(pf.get('Yappy', 0))],
        ['Tarjeta',             '', '', money(pf.get('Tarjeta', 0))],
        ['Otros',               '', '', money(pf.get('Otros', 0))],
        # 12
        ['Total Cobrado en el día', '', '', money_show(total_cobros)],
        # 13
        ['Total a Depositar en el Día', '', '', money_show(total_depositar)],
    ]

    style1 = TableStyle(
        BASE
        + sec_hdr(4)
        + col_hdr(1, 4)
        + num_right(1, 3, 2, -1)
        # filas resaltadas: Total Ventas, Total Cobrado, Total a Depositar
        + bold_bg(5,  5,  0, 3, C_LGRAY)
        + bold_bg(12, 12, 0, 3, C_LGRAY)
        + bold_bg(13, 13, 0, 3, C_MGRAY)
        # sub-encabezado Cobros (fila 6)
        + [('SPAN',        (0, 6), (3, 6)),
           ('BACKGROUND',  (0, 6), (3, 6), C_PALE),
           ('TEXTCOLOR',   (0, 6), (3, 6), C_DARK),
           ('FONTNAME',    (0, 6), (3, 6), FONT_B),
           ('LEFTPADDING', (0, 6), (3, 6), 10)]
        # filas de cobros (7-11)
        + [('BACKGROUND',  (0, 7), (3, 11), C_VLIGHT)]
        + [('SPAN', (0, r), (2, r)) for r in range(7, 12)]
        + [('LEFTPADDING', (0, 7), (0, 11), 18)]
        # spans: fila 12 y 13 (A-C label, D valor)
        + [('SPAN', (0, 12), (2, 12)),
           ('SPAN', (0, 13), (2, 13))]
    )

    t1 = Table(data1, colWidths=cw1)
    t1.setStyle(style1)
    story.append(t1)
    story.append(Spacer(1, 0.08 * inch))

    # ══════════════════════════════════════════════════════════════
    # SECCIÓN 2 — PAGOS Y GASTOS
    # Row 0: encabezado de sección
    # Row 1: encabezados de columna  (Descripción | BG | GB | Total)
    # Row 2: Pago a Proveedores       LGRAY bold
    # ══════════════════════════════════════════════════════════════
    data2 = [
        ['PAGOS Y GASTOS', '', '', ''],
        ['Descripción', 'Banco General', 'Global Bank', 'Total'],
        ['Pago a Proveedores',
         money_show(bg_egresos), money_show(gb_egresos), money_show(total_gast)],
    ]

    style2 = TableStyle(
        BASE
        + sec_hdr(4)
        + col_hdr(1, 4)
        + num_right(1, 3, 2, -1)
        + bold_bg(2, 2, 0, 3, C_LGRAY)
    )

    t2 = Table(data2, colWidths=cw1)
    t2.setStyle(style2)
    story.append(t2)
    story.append(Spacer(1, 0.15 * inch))

    # ══════════════════════════════════════════════════════════════
    # SECCIÓN 3 — MOVIMIENTOS DE BANCOS
    # Row 0: encabezado de sección
    # Row 1: encabezados de columna
    # Row 2: Saldo Anterior
    # Row 3: Depósitos
    # Row 4: Egresos
    # Row 5: Saldo Bancario al Final del Día   LGRAY bold
    # Row 6: Cheques en Circulación
    # Row 7: Saldo Bancario Real Conciliado    MGRAY bold
    # ══════════════════════════════════════════════════════════════
    data3 = [
        ['MOVIMIENTOS DE BANCOS', '', ''],
        ['Descripción', 'Banco General', 'Global Bank'],
        ['Saldo Anterior',
         money_show(bg['saldo_anterior']), money_show(gb['saldo_anterior'])],
        ['Depósitos según extracto bancario',
         money_show(bg['depositos']), money_show(gb['depositos'])],
        ['Egresos según extracto bancario',
         money(bg_egresos), money(gb_egresos)],
        ['Saldo Bancario al Final del Día',
         money_show(bg_saldo_fin), money_show(gb_saldo_fin)],
        ['Cheques en Circulación',
         money(cheques_bg), money(cheques_gb)],
        ['Saldo Bancario Real Conciliado',
         money_show(bg_conciliado), money_show(gb_conciliado)],
    ]

    style3 = TableStyle(
        BASE
        + sec_hdr(3)
        + col_hdr(1, 3)
        + num_right(1, 2, 2, -1)
        + bold_bg(5, 5, 0, 2, C_LGRAY)
        + bold_bg(7, 7, 0, 2, C_MGRAY)
    )

    t3 = Table(data3, colWidths=cw3)
    t3.setStyle(style3)
    story.append(t3)
    story.append(Spacer(1, 0.14 * inch))

    # ── Firma ─────────────────────────────────────────────────────
    story.append(Paragraph(
        'Firma: ___________________________',
        ParagraphStyle('firma', fontName=FONT, fontSize=10),
    ))
    story.append(Spacer(1, 0.06 * inch))

    # ── Leyenda ────────────────────────────────────────────────────
    story.append(Paragraph(
        'Generado automáticamente por Cierre de Caja — Grupo Panablock',
        ParagraphStyle('leyenda', fontName=FONT_I, fontSize=8,
                       textColor=C_GRAY6),
    ))

    # ── Compilar PDF ──────────────────────────────────────────────
    doc_kwargs = dict(
        pagesize=letter,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN,  bottomMargin=MARGIN,
        title='Cierre Diario de Caja — Grupo Panablock',
    )
    if output_path is None:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, **doc_kwargs)
        doc.build(story)
        buf.seek(0)
        return buf
    else:
        doc = SimpleDocTemplate(output_path, **doc_kwargs)
        doc.build(story)
        log(f"\n✔ PDF guardado en: {output_path}")
        return None
